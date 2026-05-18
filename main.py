import sys
import asyncio
import os
import datetime
import re
import uuid
import tempfile
import base64
from io import BytesIO
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image, ImageDraw

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

app = FastAPI(title="Production Site Migration Auditor")

class CompareRequest(BaseModel):
    source_url: str
    staging_url: str

def normalize_url(url: str):
    if not url.startswith(('http://', 'https://')):
        return f"https://{url}"
    return url

def clean_text_for_parity(text: str):
    return re.sub(r'\s+', '', text).lower()

def generate_side_by_side_report_to_memory(src_path, stg_path, results):
    with Image.open(src_path) as img1, Image.open(stg_path) as img2:
        img1_rgb = img1.convert('RGB')
        img2_rgb = img2.convert('RGB')
        max_h = max(img1_rgb.height, img2_rgb.height)
        gap = 20
        canvas_w = img1_rgb.width + img2_rgb.width + gap
        
        canvas = Image.new('RGB', (canvas_w, max_h), (240, 240, 240))
        canvas.paste(img1_rgb, (0, 0))
        canvas.paste(img2_rgb, (img1_rgb.width + gap, 0))
        
        draw = ImageDraw.Draw(canvas)
        for res in results:
            if res["issue"] != "PASS" and "y" in res:
                y_top = res["y"]
                height = res.get("height", 100)
                y_bottom = min(y_top + height, max_h)
                box_left = img1_rgb.width + gap
                box_right = canvas_w - 2
                draw.rectangle([box_left, y_top, box_right, y_bottom], outline="red", width=6)
        
        buffer = BytesIO()
        canvas.save(buffer, format="PNG")
        img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
        return img_str

def create_consolidated_report_to_memory(results_by_viewport):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Full Migration Audit"
    ws2 = wb.create_sheet(title="Concerns & Failures")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    
    headers = ["VIEWPORT", "COMPONENT NAME", "ISSUE CATEGORIES", "OBSERVED DISCREPANCIES", "CONTENT CHECK", "REQUIRED FIX"]
    
    for sheet in [ws1, ws2]:
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=text)
            cell.font, cell.fill, cell.alignment = header_font, header_fill, wrap
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 45
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 20
        
    f_idx, c_idx = 2, 2
    for viewport, results in results_by_viewport.items():
        for res in results:
            data = [viewport.upper(), res["name"], res["issue"], res["details"], res.get("content_diff", "Match"), res["fix"]]
            for col_idx, value in enumerate(data, 1):
                cell = ws1.cell(row=f_idx, column=col_idx, value=value)
                cell.alignment = wrap
                if res["issue"] == "PASS":
                    cell.fill = pass_fill
            f_idx += 1
            if res["issue"] != "PASS":
                for col_idx, value in enumerate(data, 1):
                    cell = ws2.cell(row=c_idx, column=col_idx, value=value)
                    cell.alignment, cell.fill = wrap, fail_fill
                c_idx += 1
                
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_b64 = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
    return excel_b64

@app.post("/compare")
def compare_sites(request: CompareRequest):
    all_viewports_results = {}
    visual_b64_reports = {}
    run_id = uuid.uuid4().hex[:8]
    src_url = normalize_url(request.source_url)
    stg_url = normalize_url(request.staging_url)
    
    viewports = {
        "desktop": {'width': 1920, 'height': 1080},
        "tablet": {'width': 768, 'height': 1024},
        "mobile": {'width': 375, 'height': 667}
    }
    
    discovery_script = """
    () => {
        const blocks = [];
        const candidates = document.querySelectorAll('header, footer, section, article, [class*="comp"], [class*="card"], [class*="banner"], [class*="alert"], [id*="banner"], [id*="alert"], main > div, div:has(> h1, > h2, > h3, > h4)');
        candidates.forEach(el => {
            if (el.closest('footer') && el.tagName !== 'FOOTER') return;
            const rect = el.getBoundingClientRect();
            const bodyText = el.innerText ? el.innerText.replace(/\\s+/g, ' ').trim() : "";
            const isBanner = /banner|alert|notification/i.test(el.className + el.id);
            const minHeight = isBanner ? 20 : 40;
            if (rect.height < minHeight || rect.width < 100 || !bodyText) return;
            const hasNested = Array.from(el.querySelectorAll('section, article, [class*="card"], [class*="comp"]')).length > 0;
            if (hasNested && el.tagName !== 'HEADER' && el.tagName !== 'FOOTER') return;
            const headline = el.querySelector('h1, h2, h3, h4, h5, .title')?.innerText?.trim();
            let label = headline || (bodyText.substring(0, 45) + "...");
            if (el.tagName === 'HEADER') label = "GLOBAL HEADER";
            if (el.tagName === 'FOOTER') label = "GLOBAL FOOTER";
            if (isBanner && el.tagName !== 'HEADER') label = "BANNER/ALERT: " + label;
            blocks.push({
                name: label,
                text: bodyText,
                y: Math.round(rect.top + window.scrollY),
                height: Math.round(rect.height),
                assets: el.querySelectorAll('img, svg, picture').length
            });
        });
        return blocks.sort((a,b) => a.y - b.y);
    }
    """
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--disable-gpu"])
            for name, size in viewports.items():
                context = browser.new_context(viewport=size, user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) MigrationAuditBot/1.0")
                page1 = context.new_page()
                page2 = context.new_page()
                tmpdir = tempfile.gettempdir()
                src_img = os.path.join(tmpdir, f"src_{name}_{run_id}.png")
                stg_img = os.path.join(tmpdir, f"stg_{name}_{run_id}.png")
                
                try:
                    for pg, url, path in [(page1, src_url, src_img), (page2, stg_url, stg_img)]:
                        pg.goto(url, wait_until="networkidle", timeout=60000)
                        pg.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                        pg.wait_for_timeout(1500)
                        pg.evaluate("window.scrollTo(0, 0)")
                        pg.wait_for_timeout(500)
                        pg.screenshot(path=path, full_page=True)
                        
                    src_data = page1.evaluate(discovery_script)
                    stg_data = page2.evaluate(discovery_script)
                    results = []
                    
                    for src in src_data:
                        match = next((s for s in stg_data if s['name'] == src['name'] or (len(src['text']) > 30 and src['text'][:30] in s['text'])), None)
                        if not match:
                            results.append({
                                "name": src['name'], "issue": "CRITICAL: Missing", 
                                "details": "Not found in Staging", "content_diff": "N/A", 
                                "fix": "Re-import", "y": src['y'], "height": src['height']
                            })
                            continue
                        
                        issues, details = [], []
                        if clean_text_for_parity(src['text']) != clean_text_for_parity(match['text']):
                            issues.append("Content Mismatch")
                            details.append("Text drift detected")
                        if src['assets'] != match['assets']:
                            issues.append("Asset Discrepancy")
                            details.append(f"Image discrepancy (Src: {src['assets']} vs Stg: {match['assets']})")
                        
                        if not issues:
                            results.append({"name": src['name'], "issue": "PASS", "details": "OK", "content_diff": "Match", "fix": "N/A"})
                        else:
                            results.append({
                                "name": src['name'], "issue": ", ".join(issues), 
                                "details": " | ".join(details), "content_diff": "Diff detected", 
                                "fix": "Review layout / implementation", "y": match['y'], "height": match['height']
                            })
                    
                    img_b64_string = generate_side_by_side_report_to_memory(src_img, stg_img, results)
                    visual_b64_reports[name] = img_b64_string
                    all_viewports_results[name] = results
                    
                finally:
                    for f in [src_img, stg_img]:
                        if os.path.exists(f):
                            try: os.remove(f)
                            except OSError: pass
                    context.close()
            browser.close()
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Migration Engine Failed: {str(e)}")
        
    excel_b64 = create_consolidated_report_to_memory(all_viewports_results)
    
    return {
        "excel_filename": f"Migration_Audit_Report_{run_id}.xlsx",
        "tabular_report_b64": excel_b64, 
        "visual_reports_b64": visual_b64_reports
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
